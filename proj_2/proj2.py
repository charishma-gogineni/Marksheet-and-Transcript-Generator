import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from fpdf import FPDF
import pandas as pd
import datetime
import re
import shutil
from flask import Flask,render_template, request, redirect

os.system('cls')

current_dir =  os.getcwd()


'''grade={"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"DD*":4,"F":0,"F*":0,"I":0} #mapping letter grades to respective numeric values

def generate_marksheet(): #function to create roll number wise .xlsx files and semester sheets in each one
    os.mkdir('output')      
    with open('grades.csv', 'r') as data:
        reader = csv.reader(data)
        for row in reader:
            if row[0]=="Roll":
                continue
            f= Workbook()       #opening workbook
            sem=['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade']
            Overall=f.active       #making an active sheet
            Overall.title = 'Overall'
            r=[['Roll No.'],['Name of Student'],['Discipline'],['Semester No.'],['Semester wise Credit Taken'],['SPI'],['Total Credits Taken'],['CPI']]
            for i in r:
                Overall.append(i)       #creating fields in the 'Overall' sheet
            
            for j in range(1,9):
                s=f.create_sheet(index=j, title='Sem{}'.format(j))
                s.append(sem)       #creating sem sheets
            
            if row[0]=="0401ME11":      #handling exceptional case for 0401ME11
                s=f.create_sheet(index=9, title='Sem10')
                s.append(sem)
                
            f.save("output//{}.xlsx".format(row[0]))
              
        data.close()
    return

def populate(sem, rollno):      #function to fill in the details in each sem sheet
    with open('grades.csv', 'r') as data:
        reader = csv.reader(data)
        b=2
        f= load_workbook("output//{}.xlsx".format(rollno[0]))      #loading already created workbook for the given roll no.
        
        sheet=f[sem]        #making the required sheet active
        credits_ncleared=0
        for r in reader:
            if r[0]=="Roll":
                continue
            
            if r[1]==sem[3:] and r[0]==rollno[0]:
                sheet.cell(row=b,column=1).value=b-1
                sheet.cell(row=b, column=2).value=r[2]
                sheet.cell(row=b,column=5).value=r[3]
                sheet.cell(row=b,column=6).value=r[5]
                sheet.cell(row=b,column=7).value=r[4].strip()
                if r[4].strip()=='F*' or r[4].strip()=='F' or r[4].strip()=='I':
                    credits_ncleared+=int(sheet.cell(row=b,column=5).value)

                b+=1
        sheet.cell(row=1, column=8).value=credits_ncleared

        f.save("output//{}.xlsx".format(rollno[0]))
        data.close()

    with open('subjects_master.csv', 'r') as data:
        reader = csv.reader(data)
        reader=list(reader)
        f= load_workbook("output//{}.xlsx".format(rollno[0]))
        sheet=f[sem]
        lst={}
        for b in range(2,sheet.max_row+1):
            lst[b]=sheet.cell(row=b,column=2).value
        for key in lst:
            for r in reader:

                if r[0]==lst[key]:
                    sheet.cell(row=key, column=3).value=r[1]
                    sheet.cell(row=key, column=4).value=r[2]

        f.save("output//{}.xlsx".format(rollno[0]))
        data.close()
    return

def overall(row):       #function to fill in data in the Overall page
    f = load_workbook("output//{}.xlsx".format(row[0]))
    sheet=f['Overall']
    sheet.cell(row=1,column=2).value=row[0]
    sheet.cell(row=2,column=2).value=row[1]
    sheet.cell(row=3,column=2).value=row[0][4:6]
    for i in range(2,10):
        sheet.cell(row=4,column=i).value=i-1
    if row[0]=='0401ME11':
        sheet.cell(row=4,column=10).value=10
    cwav=0
    for i in range(1,9):
        temp=f["Sem{}".format(i)]
        sum=wav=0
        for j in range(2,temp.max_row+1):
            cred = int(temp.cell(row=j, column=5).value)
            sum+=cred
            wav+=(grade[temp.cell(row=j,column=7).value]*cred)
        if(sum): spi=wav/sum        #spi calculation
        else: spi=0
        sheet.cell(row=5,column=i+1).value=sum      #number of credits per sem
        sheet.cell(row=6,column=i+1).value=round(spi,2) #spi 
        prev_cred=0
        if(i!=1): prev_cred=int(sheet.cell(row=7,column=i).value)       
        sheet.cell(row=7,column=i+1).value=prev_cred+sum        #cumulative credits
        cwav+=wav
        if(prev_cred==sum and sum==0): cpi=0
        else: cpi = round(cwav/(prev_cred+sum), 2)      #cpi calculation
        sheet.cell(row=8,column=i+1).value=cpi

    if row[0]=='0401ME11':
        sum=wav=0
        temp=f["Sem10"]
        for j in range(2,temp.max_row+1):
            cred = int(temp.cell(row=j, column=5).value)
            sum+=cred
            wav+=(grade[temp.cell(row=j,column=7).value]*cred)

        if(sum): spi=wav/sum        #spi calculation in exception case
        else: spi=0
        sheet.cell(row=5,column=10).value=sum      #number of credits per sem in exception case
        sheet.cell(row=6,column=10).value=round(spi,2)      #spi in exception case
        prev_cred=int(sheet.cell(row=7,column=9).value)       
        sheet.cell(row=7,column=10).value=prev_cred+sum        #cumulative credits in exception case
        cwav+=wav
        if(prev_cred==sum and sum==0): cpi=0
        else: cpi = round(cwav/(prev_cred+sum), 2)      #cpi calculation in exception case
        sheet.cell(row=8,column=10).value=cpi       

    f.save("output//{}.xlsx".format(row[0]))

generate_marksheet()        #calling function to create workbooks and sheets roll no. wise

os.system('cls')
#print("Sheets generated.")

with open('names-roll.csv','r') as data:        #calling populate and overall functions after handling exception cases
    reader=csv.reader(data)
    for r in reader:
        if(r[0]=='Roll'): continue
        for p in range(1,9):
            st="Sem"+str(p)
            populate(st, r)
        if r[0]=='0401ME11':
            populate("Sem10",r)
        overall(r)
        print(r[0], "done.")'''


# pdf_w=420
# pdf_h=297

class PDF(FPDF):
    def header(self):
        self.rect(10, 10, 400, 277)
        self.rect(10, 10, 53, 30)
        self.rect(359, 10, 51, 30)
        self.rect(10, 10, 400, 30)
        self.rect(68, 42.5, 285, 10)
        self.image('iitp_logo_PROJ2-01.png', 15, 11, 40, 25)
        self.image('iitp_logo_PROJ2-01.png', 365, 11, 40, 25)
        self.image('iitp_proj2-01.png', 90, 11, 230, 25)
        self.line(10,108, 410, 108)
        self.line(10,170, 410, 170)
        self.line(10,230, 410, 230)


course={"CS":"Computer Science and engineering","EE":"Electrical Engineering","ME":"Mechanical Engineering"}


def generate_transcript(roll):
    f=load_workbook("output\\{}.xlsx".format(roll))
    s=f['Overall']
    sem_count=len(f.sheetnames)-1

    #personal info
    pdf= PDF('L', 'mm', 'A3')
    pdf.add_page()
    pdf.set_font('Arial', 'B', 9)
    pdf.set_y(30)
    pdf.set_x(70)
    pdf.cell(100, 30, 'Roll No: {}'.format(s.cell(row=1, column=2).value), align='L')
    pdf.set_x(180)
    pdf.cell(100, 30, 'Name: {}'.format(s.cell(row=2, column=2).value), align='L')
    pdf.set_x(280)
    v=s.cell(row=1, column=2).value
    pdf.cell(100, 30, 'Year of Admission: 20{}'.format(v[0]+v[1]), align='L', ln=1)
    pdf.set_y(35)
    pdf.set_x(70)
    pdf.cell(100, 30, 'Programme: Bachelor of Technology', align='L')
    pdf.set_x(180)
    roll=s.cell(row=1, column=2).value
    pdf.cell(100, 30, 'Course: {}'.format(course[roll[4]+roll[5]]), align='L')

    pdf.set_x(10)
    pdf.set_y(58)


    if sem_count>=1:
        i=1
        pdf.ln(-3)
        pdf.set_x(10)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(10)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(10)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(10)

        t_x=10
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')


    if sem_count>=2:
        i=2
        pdf.ln(-41.5)
        pdf.set_x(155)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(155)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(155)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(155)

        t_x=155
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(155, 96, 100, 5)

    if sem_count>=3:
        i=3
        pdf.ln(-41)
        pdf.set_x(290)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(290)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(290)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(290)

        t_x=290
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(290, 88, 100, 5)

    if sem_count>=4:
        i=4
        pdf.ln(25)
        pdf.set_x(10)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(10)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(10)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(10)

        t_x=10
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(10, 154, 100, 5)

    if sem_count>=5:
        i=5
        pdf.ln(-41.5)
        pdf.set_x(155)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(155)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(155)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(155)
        t_x=155
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(155, 149.5, 100, 5)

    if sem_count>=6:
        i=6
        pdf.ln(-38)
        pdf.set_x(290)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(290)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(290)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(290)

        t_x=290
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(290, 148.5, 100, 5)

    if sem_count>=7:
        i=7
        pdf.ln(23)
        pdf.set_x(10)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(10)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(10)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(10)

        t_x=10
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(10, 208.5, 100, 5)

    if sem_count>=8:
        i=8
        pdf.ln(-39)
        pdf.set_x(155)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i), ln=1)

        pdf.set_x(155)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)
        
        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(155)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(155)

        t_x=155
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(155, 198.5, 100, 5)


    if sem_count>=9:
        i=9
        pdf.ln(-30)
        pdf.set_x(290)
        pdf.set_font('Arial', 'BU', 9)
        pdf.cell(20,5,"Semester {}".format(i+1), ln=1)

        pdf.set_x(290)
        d=pd.read_excel("output\\{}.xlsx".format(roll), sheet_name=i)

        d.drop(['Sl No.', 'Subject Type'], axis=1, inplace=True)
        l=list(d.columns)

        #table created
        pdf.set_font('Arial', 'B', 7)
        for h in l[:-1]:
            if h=='Subject No.':
                pdf.cell(14, 4, h, 1, 0, 'C')
            elif h=='Subject Name':
                pdf.cell(70, 4, h, 1, 0, 'C')
            elif h =='L-T-P':
                pdf.cell(13, 4, h, 1, 0, 'C')
            elif h =="Credit":
                pdf.cell(10, 4, h, 1, 0, 'C')
            else:
                continue
        pdf.cell(10, 4, l[4], 1,1,'C')

        pdf.set_x(290)
        pdf.set_font('Arial', '', 8)
        for r in range(0, len(d)):
            for c, col_name in enumerate(l):
                if c!=len(l)-2:
                    if col_name == 'Subject No.':
                        pdf.cell(14, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Subject Name":
                        pdf.cell(70, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name =="L-T-P":
                        pdf.cell(13, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                    if col_name == "Credit":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
                else:
                    if col_name == "Grade":
                        pdf.cell(10, 4, str(d['%s'%(col_name)].iloc[r]),1,0,'C')
            pdf.cell(35,4,"",ln=1)
            pdf.set_x(290)

        t_x=290
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(30, 5, "Credits Taken:  {}".format(s.cell(row=5, column=i+1).value), 'L,B,T', 0,'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(30, 5, "Credits Cleared:  {}".format(s.cell(row=5, column=i+1).value-int(l[-1])), 'B,T', 0, 'L')
        t_x+=30
        pdf.set_x(t_x)
        pdf.cell(20, 5, "SPI:  {}".format(s.cell(row=6, column=i+1).value), 'B,T', 0, 'L')
        t_x+=20
        pdf.set_x(t_x)
        pdf.cell(30, 5, "CPI:  {}".format(s.cell(row=8, column=i+1).value), 'R,B,T', 0,'L')
        #pdf.rect(290, 198.5, 100, 5)

    e=datetime.datetime.now()
    
    pdf.set_xy(x=25,y=260)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(35, 10, "Date Generated: {}, {}".format(e.strftime("%d %b, %Y"), e.strftime("%H:%M")), align='L')
    if seal_c==0:
        pass
    else:
        #print(seal_n)
        #print(type(seal_n))
        pdf.image('uploads\\{}'.format(str(seal_n)), 190, 240, 40, 40)
    pdf.set_x(330)
    pdf.cell(35, 2, "___________________________", 'L', ln=1)
    pdf.set_x(330)
    pdf.cell(35, 10, "Assistant Registrar (Academic)", align='L', ln=1)
    if sign_c==0:
        pass
    else:
        pdf.image('uploads\\{}'.format(str(sign_n)), 340, 235, 40, 25)

    pdf.output('pdfs\\{}.pdf'.format(roll))
    return


seal_c = 0
sign_c = 0
seal_n = ""
sign_n = ""

def generate_range(roll_1, roll_2):
    # if roll_1 == roll_2:
    #     generate_transcript(roll_1)
    #     return
    l = []
    patt = roll_1[:6]
    s = int(roll_1[6:8])
    e = int(roll_2[6:8])
    if e>=s:
        for x in range (s,e+1):
            with open('names-roll.csv', 'r') as file:
                reader = csv.reader(file)
                c = 0
                for roll in reader:
                    x = str(x)
                    if roll[0]== patt + x.zfill(2):
                        c+=1
                        generate_transcript(roll[0])
                        continue
            if c==0:
                roll_not_present = patt + str(x)
                l.append(roll_not_present)
    if s>e:
        for x in range (e,s+1):
            with open('names-roll.csv', 'r') as file:
                reader = csv.reader(file)
                c = 0
                for roll in reader:
                    if roll[0]== patt + str(x):
                        c+=1
                        generate_transcript(roll[0])
                        continue
            if c==0:
                roll_not_present = patt + str(x)
                l.append(roll_not_present)
    return l

def generate_all_transcripts():
    with open('names-roll.csv', 'r') as traverse:
        roll_numbers = csv.reader(traverse)
        for roll in roll_numbers:
            if roll[0]=="Roll":
                continue
            temp_roll= roll[0]
            if temp_roll[2]=='0' and temp_roll[3]=='1':
                generate_transcript(temp_roll)
                # print(temp_roll, "done")
    return



def function_to_check_range(r):
    r = r.upper()
    regex = r"([0-9]{4}[A-Z]{2}[0-9]{2})([-])([0-9]{4}[A-Z]{2}[0-9]{2})"
    if re.match(regex,r):
        match = re.search(regex,r)
        start_roll = match.group(1)
        end_roll = match.group(3)
        # print(start_roll, end_roll)
        for x in range(0,6):
            if start_roll[x]!=end_roll[x]:
                return False
            else:
                pass
        return True
    return False

def clear_things():
    try:
        shutil.rmtree(fr'{current_dir}\pdfs')
    except FileNotFoundError:
        pass
    os.mkdir('pdfs')
    return

app = Flask(__name__, template_folder = "templates")
app.config['UPLOAD_PATH'] = fr'{current_dir}\uploads'
@app.route("/")
def form():
    
    return render_template("proj2.html")

@app.route("/data", methods = ['POST','GET'])
def run():
    clear_things()
    # seal_c = 0
    # sign_c = 0
    if request.method=='POST':
        if request.files:
            try:
                seal = request.files['seal']
                #print(seal)
                seal.save(os.path.join(app.config['UPLOAD_PATH'],seal.filename))
                global seal_n
                seal_n=seal.filename
                global seal_c
                seal_c=1
                #print(seal_n)
            except FileNotFoundError:
                pass
            try:
                sign = request.files['sign']
                sign.save(os.path.join(app.config['UPLOAD_PATH'],sign.filename))
                global sign_n
                sign_n=sign.filename
                global sign_c
                sign_c=1
            except FileNotFoundError:
                pass
    if seal_c ==0:
        print("no seal")
    if sign_c==0:
        print("no sign")
    if request.method =='POST':
        try:
            generate_all = request.form['transcripts']
        except KeyError:
            return "<h2>Please provide complete information</h2>"
        if generate_all == "no":
            try:
                r = request.form['range']
                rolls = function_to_check_range(r)
                if rolls == False:
                    return "<h2>Please enter a valid range!</h2>"
                else:
                    r = r.upper()
                    regex = r"([0-9]{4}[A-Z]{2}[0-9]{2})([-])([0-9]{4}[A-Z]{2}[0-9]{2})"
                    match = re.search(regex,r)
                    if match != None:
                        start_roll = match.group(1)
                        end_roll = match.group(3)
                    # start_roll 
                    lists = generate_range(start_roll, end_roll)
                    print(lists)
                    # length_lists = len(lists)
                    return render_template("pass.html", lists = lists)
                    if length_lists == 0:
                        return "<h2>All required transcripts Generated</h2>"
                    else:
                        return "Required transcripts generated except {{l}}"
            except KeyError:
                return "<h2>Please provide complete information</h2>"
            # list_un = generate_range(roll_1, roll_2)
        elif generate_all =="yes":
            generate_all_transcripts()
            return "<h2>All files generated!</h2>"
if __name__ =='__main__':
    app.run(debug = True)