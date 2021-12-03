from flask import Flask,render_template, request, redirect
import csv
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
# from openpyxl.drawing.image import Image
from PIL import Image
import shutil
os.system('cls')

current_dir = os.getcwd()

def clear_things():
    os.chdir(fr'{current_dir}\sample_output')
    try:
        shutil.rmtree(fr'{current_dir}\sample_output\marksheets')
    except FileNotFoundError:
        pass


os.chdir(current_dir)
app = Flask(__name__,template_folder = "templates")

def check_ANSWER(l,address):
    x=0
    
    os.chdir(fr'{address}\uploads')
    with open(str(filename_responses),'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[6]=="ANSWER":
                for i in range(7,100):
                    try:
                        if row[i] != None:
                            x = x + 1
                            l.append(row[i])
                    except IndexError:
                        break
                return
            else:
                continue
        if x==0:
            return
def execute(current_ans_,wrong_ans_):
    address = current_dir
    correct_m = current_ans_
    wrong_m = wrong_ans_
    l = [] 
    check_ANSWER(l,address)
    noq = len(l)
    # print(l)

    if noq == 0:
        return "No roll number with ANSWER is present, Cannot Process!"
        exit()
    thin = Side(border_style = "thin", color = "000000")
    try:
        os.chdir(fr'{address}\sample_output\marksheets')
    except FileNotFoundError:
        os.mkdir(fr'{address}\sample_output\marksheets')

    os.chdir(fr'{address}\uploads')
    with open(str(filename_master_roll), 'r') as file:
        
        reader = csv.reader(file)
        for row in reader:
            if row[0] == 'roll':
                continue
            temp = []
            os.chdir(fr'{address}\uploads')
            with open(str(filename_responses),'r') as f:
                read = csv.reader(f)
                result = []
                for r in read:
                    if r[0] == "Timestamp":
                        try:
                            os.chdir(fr'{address}\sample_output\marksheets')
                            with open('concise_marksheet.csv','x',newline='') as concise:
                                    writer = csv.writer(concise)
                                    append_it = []
                                    append_it.extend(r[:6])
                                    append_it.extend(['Score_After_Negative'])
                                    append_it.extend(r[6:]) 
                                    append_it.extend(['statusAns'])
                                    writer.writerow(append_it)
                                    continue
                        except FileExistsError:
                            continue
                        
                    if row[0].upper() != r[6].upper():
                        continue
                    temp.append(row[0].upper())
                    correct_count=0
                    wrong_count=0
                    for i in range(6+1,6+noq+1):
                        result.append(r[i])
                        if l[i-7] == r[i]:
                            correct_count = correct_count + 1
                        elif r[i]==None:
                            pass
                        else:
                            wrong_count = wrong_count + 1
                    attempted = correct_count + wrong_count
                    
                    from openpyxl import Workbook
                    wb = Workbook()
                    sheet = wb.active
                    sheet.column_dimensions['A'].width = 17
                    sheet.column_dimensions['B'].width = 17
                    sheet.column_dimensions['C'].width = 17
                    sheet.column_dimensions['D'].width = 17
                    sheet.column_dimensions['E'].width = 17
                    width = 85
                    height = 60
                    img = openpyxl.drawing.image.Image(fr'{address}\iitp_logo.png')
                    sheet.add_image(img, "A1")
                    sheet.merge_cells("A5:E5")
                    a1 = sheet['A5']
                    a1.value = "Mark Sheet"
                    a1.alignment = Alignment(horizontal="center",vertical="center")
                    a1.font = Font(name = "Century", size = 18, bold = True)
                    
                    for i in range (9,13):
                        for j in range(1,6):

                            x1 = sheet.cell(row = i, column=j)
                            # x1.border = Border(top = thin, left = thin, right = thin, bottom=thin)
                            if i==9 and j==2:
                                x1.value = "Right"
                            if i==9 and j==3:
                                x1.value = "Wrong"
                            if i==9 and j==4:
                                x1.value = "Not Attempted"
                            if i==9 and j==5:
                                x1.value = "Max"

                            if i==10 and j==1:
                                x1.value = "No."
                            if i==10 and j==2:
                                x1.value = correct_count
                            if i==10 and j==3:
                                x1.value = wrong_count
                            if i==10 and j==4:
                                x1.value = noq - correct_count - wrong_count
                            if i==10 and j==5:
                                x1.value = noq

                            if i==11 and j==1:
                                x1.value = "Marking"
                            if i==11 and j==2:
                                x1.value = correct_m
                            if i==11 and j==3:
                                x1.value = wrong_m
                            if i==11 and j==4:
                                x1.value = 0

                            if i==12 and j==1:
                                x1.value = "Total"
                            if i==12 and j==2:
                                x1.value = correct_count*correct_m
                            if i==12 and j==3:
                                x1.value = wrong_count*wrong_m
                            if i==12 and j==5:
                                sum = correct_count*correct_m + wrong_count*wrong_m
                                max_total = noq*correct_m
                                sum = round(sum, 3)
                                max_total = round(max_total,3)
                                x1.value = f'{sum}/{max_total}'
                            x1.font = Font(size = 12, name = 'Century', bold = True)
                    for i in range(10,13):
                        for j in range(2,6):
                            x1 = sheet.cell(row = i, column = j)
                            if x1.value==None:
                                continue
                            if j==2:
                                x1.font = Font(name = "Century", size = 12, bold = False, color = "008000")
                            elif j==3:
                                x1.font = Font(name = "Century", size = 12, bold = False, color = "FF0000")
                            elif i==12 and j==5:
                                x1.font = Font(name = "Century", size = 12, bold = False, color = "0000FF")                       
                            else:
                                x1.font = Font(name = "Century", size = 12, bold = False)

                    sheet["A6"] = 'Name:'
                    sheet['A6'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet['B6'] = row[1]
                    sheet['B6'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["A7"] = "Roll Number:"
                    sheet['A7'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["B7"] = row[0].upper()
                    sheet['B7'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["D6"] = "Exam"
                    sheet['D6'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["E6"] = "Quiz"
                    sheet['E6'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["D7"] = "Attendance"
                    sheet['D7'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["E7"] = "Present"
                    sheet['E7'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["A15"] = "Student Ans"
                    sheet['A15'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["B15"] = "Correct Ans"
                    sheet['B15'].font = Font(size = 12, name = 'Century', bold = True)

                    sheet["A15"].border = Border(top = thin, left = thin, right = thin, bottom=thin)
                    sheet["B15"].border = Border(top = thin, left = thin, right = thin, bottom=thin)

                    for i in range(0,noq):
                        x1 = sheet.cell(row = 16 + i, column = 1)
                        x1.value = result[i]
                        x2 = sheet.cell(row = 16 + i, column = 2)
                        x2.value = l[i]
                        x2.font = Font(size = 12, name = 'Century', color = "0000FF")
                        
                        x1.border = Border(top = thin, left = thin, right = thin, bottom=thin)
                        x2.border = Border(top = thin, left = thin, right = thin, bottom=thin)

                        
                        if x1.value == x2.value:
                            x1.font = Font(size = 12, name = 'Century', color = "008000")
                        elif x1.value == None:
                            pass
                        else:
                            x1.font = Font(size = 12, name = 'Century', color = "FF0000")
                    try:
                        os.chdir(fr'{address}\sample_output\marksheets')
                    except FileNotFoundError:
                        os.mkdir(fr'{address}\sample_output\marksheets')
                        os.chdir(fr'{address}\sample_output\marksheets')
                    wb.save(filename=f'{row[0]}.xlsx')
                    os.chdir(fr'{address}\sample_output\marksheets')
                    with open('concise_marksheet.csv','a',newline='') as concise:
                        writer = csv.writer(concise)
                        append_it = []
                        append_it.extend(r[:6])
                        append_it.extend([f'{sum}/{max_total}'])
                        append_it.extend(r[6:])
                        x = [correct_count, wrong_count, noq-correct_count-wrong_count]
                        append_it.append(x)
                        writer.writerow(append_it)



                if row[0] not in temp:
                    from openpyxl import Workbook
                    wb = Workbook()
                    sheet = wb.active
                    sheet.column_dimensions['A'].width = 17
                    sheet.column_dimensions['B'].width = 17
                    sheet.column_dimensions['C'].width = 17
                    sheet.column_dimensions['D'].width = 17
                    sheet.column_dimensions['E'].width = 17
                    width = 85
                    height = 60
                    img = openpyxl.drawing.image.Image(fr'{address}\iitp_logo.png')
                    sheet.add_image(img, "A1")
                    sheet.merge_cells("A5:E5")
                    a1 = sheet['A5']
                    a1.value = "Mark Sheet"
                    a1.alignment = Alignment(horizontal="center",vertical="center")
                    a1.font = Font(name = "Century", size = 18, bold = True)

                    sheet["A6"] = 'Name:'
                    sheet['A6'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet['B6'] = row[1]
                    sheet['B6'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["A7"] = "Roll Number:"
                    sheet['A7'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["B7"] = row[0].upper()
                    sheet['B7'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["D6"] = "Exam"
                    sheet['D6'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["E6"] = "Quiz"
                    sheet['E6'].font = Font(size = 12, name = 'Century', bold = True)
                    sheet["D7"] = "Attendance"
                    sheet['D7'].font = Font(size = 12, name = 'Century', bold = False)
                    sheet["E7"] = "Absent"
                    sheet['E7'].font = Font(size = 12, name = 'Century', bold = True)
                    os.chdir(fr'{address}\sample_output\marksheets')
                    wb.save(filename=f'{row[0]}.xlsx')

    return "Done"

cur = current_dir
def send_mail_exe(eid, a,b):
            email_user =  'cs384.python4proj@gmail.com' #'cs381.email@gmail.com'
            email_send_to = eid #receiver email
            subject = 'Python quiz marks released' #subject of the mail

            msg = MIMEMultipart() 
            msg['From'] = email_user
            msg['To'] = email_send_to
            msg['Subject'] = subject

            body = f"Hey {b}({a})! Here is your marksheet of the quiz." 
            msg.attach(MIMEText(body,'plain')) 
            os.chdir(fr'{cur}\sample_output\marksheets')
            filename = f'{a}.xlsx' 
            attachment = open(filename,'rb')
            part = MIMEBase('application','octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',"attachment; filename= "+filename)

            msg.attach(part) 
            text = msg.as_string()

            server = smtplib.SMTP('smtp.gmail.com',587)
            server.starttls()
            server.login(email_user,'Pythonitis') 

            server.sendmail(email_user,email_send_to,text) 
            server.quit()
            return
def send_concise(sending_email):
    email_user = 'cs384.python4proj@gmail.com'
    email_send_to = sending_email
    subject = 'Concise Marksheet of Quiz'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send_to
    msg['Subject'] = subject

    body = "The concise marksheet is generated and attached with the mail. Thank you!"
    msg.attach(MIMEText(body,'plain'))

    os.chdir(fr'{cur}\sample_output\marksheets')
    filename = 'concise_marksheet.csv'
    attachment = open(filename,'rb')
    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',"attachment; filename= "+filename)

    msg.attach(part)
    text = msg.as_string()


    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(email_user,'Pythonitis')


    server.sendmail(email_user,email_send_to,text)
    server.quit()
    return
def send_mail():
    os.chdir(fr'{cur}\uploads')
    with open(str(filename_master_roll),'r') as f:
        reader = csv.reader(f)
        for row in reader:
            eid = None
            if row[0]=='roll' or row[0]=="ANSWER":
                continue
            else:
                os.chdir(fr'{cur}\uploads')
                with open(str(filename_responses),'r') as file:
                    read = csv.reader(file)
                    for r in read:
                        if r[6].upper()==row[0].upper():
                            eid = r[1]
                            eid2 = r[4]
                            break
                        else:
                            continue
            a = row[0]
            b = row[1]
            send_mail_exe(eid, a, b)
            send_mail_exe(eid2, a,b)
            
    return
@app.route("/")
def form():
    return render_template("base.html")

app.config['UPLOAD_PATH'] = fr'{current_dir}\uploads'

filename_master_roll = ""
filename_responses = ""
@app.route("/data", methods=['POST','GET'])
def getvalue():
    clear_things()
    master_roll = None
    responses = None
    
    try:
        correct_ans_ = request.form['correct_ans']
        wrong_ans_ = request.form['wrong_ans']
        generate_roll = request.form["Generate_roll_no_wise_Marksheet"]
        generate_concise = request.form["Generate_concise_Marksheet"]
        sending_email = request.form['email']
    except KeyError:
        return "<h2>Information missing. Please provide all the details!<h2>"
    if correct_ans_=='' or wrong_ans_=="":
        return '<h2>Please Enter data completely!</h2>'
    email = request.form['email']
    if request.method=='POST':
        if request.files:
            master_roll = request.files['master_roll']
            responses = request.files['responses']
            if master_roll=='' or responses=='':
                return 'One or more files were not uploaded!'
            else:
                try:
                    master_roll.save(os.path.join(app.config['UPLOAD_PATH'],master_roll.filename))
                    responses.save(os.path.join(app.config['UPLOAD_PATH'],responses.filename))
                    global filename_master_roll 
                    filename_master_roll = master_roll.filename

                    global filename_responses 
                    filename_responses = responses.filename
                except FileNotFoundError:
                    return "<h2>One or more files were not uploaded!</h2>"
                s = execute(float(correct_ans_), float(wrong_ans_))
                if s == "No roll number with ANSWER is present, Cannot Process!":
                    return "<h2>No roll number with ANSWER is present, Cannot Process!</h2>"
                
                if generate_roll == "yes_gr":
                    send_mail()
                if generate_concise == "yes_gc":
                    send_concise(sending_email)
    return render_template('pass.html',e = email,gr = generate_roll, gc =  generate_concise)
    
if __name__ =='__main__':
    app.run(debug = True)